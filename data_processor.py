"""
Módulo de procesamiento de datos para evidencias de cobranzas
Maneja la sanitización de campos y generación de archivos de evidencias
"""
import pandas as pd
import numpy as np
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import numbers
from typing import Dict, List, Tuple, Optional


class DataProcessor:
    """Procesador de datos para generación de evidencias de gestión"""
    
    def __init__(self, log_callback=None):
        """
        Inicializa el procesador de datos
        
        Args:
            log_callback: Función para enviar mensajes de log a la interfaz
        """
        self.log_callback = log_callback
        
        # Mapeo de nombres de campos para sanitización
        self.field_mappings = {
            'cuenta': ['cuenta', 'CUENTA', 'Cuenta'],
            'nombre': ['nombre', 'NOMBRE', 'nombres', 'NOMBRES', 'contacto', 'CONTACTO', 
                      'nombre completo', 'NOMBRE COMPLETO', 'nombre_completo', 'NOMBRE_COMPLETO'],
            'dni': ['dni', 'DNI', 'documento', 'DOCUMENTO', 'Dni', 'Documento'],
            'gestion_efectiva': ['gestion efectiva', 'GESTION EFECTIVA', 'gestión efectiva', 
                                'GESTIÓN EFECTIVA', 'gestion_efectiva', 'GESTION_EFECTIVA'],
            'telefono': ['telefono', 'TELEFONO', 'teléfono', 'TELÉFONO', 'celular', 
                        'CELULAR', 'Telefono', 'Celular'],
            'tipo_gestion': ['tipo de gestion', 'TIPO DE GESTION', 'tipo_gestion', 'TIPO_GESTION',
                           'tipo de gestión', 'TIPO DE GESTIÓN'],
            'numero_credito': ['numero de credito', 'NUMERO DE CREDITO', 'número de crédito',
                             'NÚMERO DE CRÉDITO', 'numero_credito', 'NUMERO_CREDITO'],
            'ruta': ['ruta', 'RUTA', 'Ruta'],
            'nombre_completo_audio': ['nombre_completo', 'NOMBRE_COMPLETO', 'nombre completo']
        }
    
    def log(self, message: str):
        """Envía un mensaje de log a la interfaz"""
        if self.log_callback:
            self.log_callback(message)
    
    def save_excel_formatted(self, df: pd.DataFrame, excel_path: Path):
        """
        Guarda un DataFrame a Excel con formato de texto para campos numéricos
        y sin valores NaN (se muestran como celdas vacías)
        
        Args:
            df: DataFrame a guardar
            excel_path: Ruta donde guardar el archivo Excel
        """
        # Crear copia para no modificar el original
        df_formatted = df.copy()
        
        # Reemplazar NaN con cadena vacía
        df_formatted = df_formatted.fillna('')
        
        # Columnas que típicamente contienen números largos
        numeric_columns = ['cuenta', 'telefono', 'celular', 'dni', 'documento',
                          'numero_credito', 'CUENTA', 'TELEFONO', 'CELULAR', 'DNI',
                          'DOCUMENTO', 'NUMERO DE CREDITO', 'numero de credito']
        
        # Convertir columnas numéricas a texto para evitar notación científica
        for col in df_formatted.columns:
            col_lower = col.lower().strip()
            # Verificar si es una columna numérica conocida o si contiene valores numéricos largos
            if col in numeric_columns or col_lower in [c.lower() for c in numeric_columns]:
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: str(int(float(x))) if x != '' and pd.notna(x) and str(x).replace('.', '').replace('-', '').isdigit() else (str(x) if x != '' else '')
                )
            else:
                # Para otras columnas, convertir números a string si son muy largos
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: str(int(float(x))) if isinstance(x, (int, float)) and not isinstance(x, bool) and pd.notna(x) and len(str(int(float(x)))) > 10 else x
                )
        
        # Guardar el archivo Excel
        df_formatted.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Abrir el archivo y aplicar formato de texto a las columnas numéricas
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Obtener índices de columnas numéricas
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        numeric_col_indices = []
        for idx, col_name in enumerate(header_row, 1):
            if col_name:
                col_lower = str(col_name).lower().strip()
                if col_name in numeric_columns or col_lower in [c.lower() for c in numeric_columns]:
                    numeric_col_indices.append(idx)
        
        # Aplicar formato de texto a las celdas de columnas numéricas
        for col_idx in numeric_col_indices:
            for row in range(2, ws.max_row + 1):  # Empezar desde la fila 2 (después del header)
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = numbers.FORMAT_TEXT
        
        wb.save(excel_path)
    
    def sanitize_dataframe(self, df: pd.DataFrame, skip_consolidados: bool = False) -> pd.DataFrame:
        """
        Sanitiza los nombres de columnas de un DataFrame
        
        Args:
            df: DataFrame a sanitizar
            skip_consolidados: Si es True, no sanitiza (para consolidados.xlsx)
            
        Returns:
            DataFrame con columnas sanitizadas
        """
        if skip_consolidados:
            # Solo quitar espacios en blanco de valores, no cambiar nombres de columnas
            df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
            return df
        
        df_copy = df.copy()
        
        # Renombrar columnas según el mapeo
        column_rename = {}
        for standard_name, variations in self.field_mappings.items():
            for col in df_copy.columns:
                if col.strip() in variations:
                    column_rename[col] = standard_name
                    break
        
        df_copy.rename(columns=column_rename, inplace=True)
        
        # Quitar espacios en blanco de los valores
        for col in df_copy.columns:
            if df_copy[col].dtype == 'object':
                df_copy[col] = df_copy[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        return df_copy
    
    def parse_gestion_efectiva(self, gestion_str: str) -> List[str]:
        """
        Parsea el campo GESTION EFECTIVA separado por comas
        
        Args:
            gestion_str: String con gestiones separadas por coma
            
        Returns:
            Lista de gestiones (IVR, SMS, CALL, GRABACION CALL)
        """
        if pd.isna(gestion_str):
            return []
        
        gestiones = [g.strip().upper() for g in str(gestion_str).split(',')]
        
        # Normalizar GRABACION CALL a CALL
        gestiones = ['CALL' if 'CALL' in g else g for g in gestiones]
        
        return list(set(gestiones))  # Eliminar duplicados
    
    def create_ivr_evidence(self, cliente_data: Dict, nuevos_datos_df: pd.DataFrame, 
                           output_folder: Path, audio_ivr_path: str) -> Tuple[bool, List[str]]:
        """
        Crea archivos de evidencia IVR para un cliente
        
        Returns:
            Tuple (success, files_created)
        """
        files_created = []
        
        try:
            cuenta = cliente_data['cuenta']
            nombre = cliente_data['nombre']
            
            # Copiar audio IVR (SIEMPRE se copia si el cliente tiene gestión IVR)
            audio_filename = f"ivr_{nombre}.mp3"
            audio_path = output_folder / audio_filename
            shutil.copy2(audio_ivr_path, audio_path)
            files_created.append(audio_filename)
            
            # Filtrar en nuevos_datos por CUENTA y GESTION_EFECTIVA = IVR
            ivr_data = nuevos_datos_df[
                (nuevos_datos_df['cuenta'] == cuenta) & 
                (nuevos_datos_df['gestion_efectiva'].str.contains('IVR', na=False))
            ].copy()
            
            if ivr_data.empty:
                self.log(f"  ⚠️ No se encontraron registros IVR en nuevos_datos para {nombre} (audio IVR copiado)")
            else:
                # Agregar columna TIPO DE GESTION
                ivr_data['TIPO DE GESTION'] = 'IVR'
                
                # Crear archivo Excel con formato de texto para campos numéricos
                excel_filename = f"{nombre}_ivr.xlsx"
                excel_path = output_folder / excel_filename
                self.save_excel_formatted(ivr_data, excel_path)
                files_created.append(excel_filename)
            
            return True, files_created
            
        except Exception as e:
            self.log(f"  ❌ Error creando evidencia IVR: {str(e)}")
            return False, files_created
    
    def create_sms_evidence(self, cliente_data: Dict, sms_df: pd.DataFrame, 
                           output_folder: Path) -> Tuple[bool, List[str]]:
        """
        Crea archivo de evidencia SMS para un cliente
        
        Returns:
            Tuple (success, files_created)
        """
        files_created = []
        
        try:
            cuenta = cliente_data['cuenta']
            nombre = cliente_data['nombre']
            
            # Filtrar en sms.xlsx por NUMERO DE CREDITO
            sms_data = sms_df[sms_df['numero_credito'] == cuenta].copy()
            
            if sms_data.empty:
                self.log(f"  ⚠️ No se encontraron registros SMS para {nombre}")
                return False, files_created
            
            # Crear archivo Excel con formato de texto para campos numéricos
            excel_filename = f"SMS_{nombre}.xlsx"
            excel_path = output_folder / excel_filename
            self.save_excel_formatted(sms_data, excel_path)
            files_created.append(excel_filename)
            
            return True, files_created
            
        except Exception as e:
            self.log(f"  ❌ Error creando evidencia SMS: {str(e)}")
            return False, files_created
    
    def create_call_evidence(self, cliente_data: Dict, nuevos_datos_df: pd.DataFrame,
                            consolidados_df: Optional[pd.DataFrame], output_folder: Path) -> Tuple[bool, List[str]]:
        """
        Crea archivos de evidencia CALL para un cliente
        
        Returns:
            Tuple (success, files_created)
        """
        files_created = []
        
        try:
            cuenta = cliente_data['cuenta']
            nombre = cliente_data['nombre']
            dni = cliente_data.get('dni', '')
            telefono = cliente_data.get('telefono', '')
            
            # Filtrar en nuevos_datos por CUENTA y GESTION_EFECTIVA = CALL
            call_data = nuevos_datos_df[
                (nuevos_datos_df['cuenta'] == cuenta) & 
                (nuevos_datos_df['gestion_efectiva'].str.contains('CALL', na=False))
            ].copy()
            
            if call_data.empty:
                self.log(f"  ⚠️ No se encontraron registros CALL en nuevos_datos para {nombre}")
                return False, files_created
            
            # Agregar columna TIPO DE GESTION
            call_data['TIPO DE GESTION'] = 'CALL'
            
            # Crear archivo Excel con formato de texto para campos numéricos
            excel_filename = f"{nombre}_gestiones.xlsx"
            excel_path = output_folder / excel_filename
            self.save_excel_formatted(call_data, excel_path)
            files_created.append(excel_filename)
            
            # Buscar audio en consolidados (OPCIONAL - solo si existe consolidados_df)
            if consolidados_df is not None:
                audio_found = False
                audio_row = None
                
                # Primero intentar buscar por DNI
                if dni:
                    audio_row = consolidados_df[consolidados_df['dni'].astype(str) == str(dni)]
                    if not audio_row.empty:
                        audio_found = True
                
                # Si no se encontró por DNI, buscar por teléfono
                if not audio_found and telefono:
                    audio_row = consolidados_df[consolidados_df['telefono'].astype(str) == str(telefono)]
                    if not audio_row.empty:
                        audio_found = True
                
                if audio_found and not audio_row.empty:
                    # Construir ruta del audio
                    ruta = str(audio_row.iloc[0]['ruta'])
                    nombre_completo_audio = str(audio_row.iloc[0]['nombre_completo'])
                    audio_source_path = f"{ruta}/{nombre_completo_audio}.mp3"
                    
                    if os.path.exists(audio_source_path):
                        # Copiar audio
                        audio_filename = f"{nombre}_{cuenta}.mp3"
                        audio_dest_path = output_folder / audio_filename
                        shutil.copy2(audio_source_path, audio_dest_path)
                        files_created.append(audio_filename)
                    else:
                        self.log(f"  ⚠️ Audio no encontrado en: {audio_source_path}")
                else:
                    self.log(f"  ⚠️ No se encontró audio CALL para {nombre} (DNI: {dni}, TEL: {telefono}) - Excel creado")
            else:
                self.log(f"  ℹ️ consolidados.xlsx no proporcionado - Solo Excel CALL creado para {nombre}")
            
            return True, files_created
            
        except Exception as e:
            self.log(f"  ❌ Error creando evidencia CALL: {str(e)}")
            return False, files_created
    
    def process_cliente(self, cliente_row: pd.Series, nuevos_datos_df: pd.DataFrame,
                       sms_df: Optional[pd.DataFrame], consolidados_df: Optional[pd.DataFrame],
                       audio_ivr_path: str, base_output_folder: Path) -> bool:
        """
        Procesa un cliente individual y crea sus archivos de evidencia
        
        Returns:
            True si se procesó exitosamente
        """
        try:
            # Extraer datos del cliente
            cuenta = cliente_row['cuenta']
            nombre = cliente_row['nombre']
            dni = cliente_row.get('dni', '')
            telefono = cliente_row.get('telefono', '')
            gestion_efectiva_str = cliente_row['gestion_efectiva']
            
            # Parsear gestiones efectivas
            gestiones = self.parse_gestion_efectiva(gestion_efectiva_str)
            
            if not gestiones:
                self.log(f"⚠️ Cliente {nombre} no tiene gestiones efectivas")
                return False
            
            # Crear carpeta del cliente
            folder_name = f"{nombre}_{cuenta}"
            cliente_folder = base_output_folder / folder_name
            cliente_folder.mkdir(parents=True, exist_ok=True)
            
            self.log(f"\n📁 Procesando: {folder_name}")
            self.log(f"  Gestiones: {', '.join(gestiones)}")
            
            cliente_data = {
                'cuenta': cuenta,
                'nombre': nombre,
                'dni': dni,
                'telefono': telefono
            }
            
            files_created_total = []
            
            # Procesar IVR
            if 'IVR' in gestiones:
                success, files = self.create_ivr_evidence(
                    cliente_data, nuevos_datos_df, cliente_folder, audio_ivr_path
                )
                if success:
                    files_created_total.extend(files)
                    self.log(f"  ✅ IVR: {', '.join(files)}")
            
            # Procesar SMS
            if 'SMS' in gestiones and sms_df is not None:
                success, files = self.create_sms_evidence(
                    cliente_data, sms_df, cliente_folder
                )
                if success:
                    files_created_total.extend(files)
                    self.log(f"  ✅ SMS: {', '.join(files)}")
            
            # Procesar CALL (consolidados_df es opcional)
            if 'CALL' in gestiones:
                success, files = self.create_call_evidence(
                    cliente_data, nuevos_datos_df, consolidados_df, cliente_folder
                )
                if success:
                    files_created_total.extend(files)
                    self.log(f"  ✅ CALL: {', '.join(files)}")
            
            self.log(f"  📊 Total archivos creados: {len(files_created_total)}")
            
            return True
            
        except Exception as e:
            self.log(f"❌ Error procesando cliente {nombre}: {str(e)}")
            return False
    
    def validate_dataframe_fields(self, df: pd.DataFrame, required_fields: List[str], 
                                  file_name: str) -> Tuple[bool, str]:
        """
        Valida que un DataFrame contenga los campos requeridos
        
        Returns:
            Tuple (valid, error_message)
        """
        missing_fields = []
        for field in required_fields:
            if field not in df.columns:
                missing_fields.append(field)
        
        if missing_fields:
            return False, f"{file_name}: Faltan campos {', '.join(missing_fields)}"
        
        return True, ""
